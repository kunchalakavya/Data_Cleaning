"""
generate_report.py
Reads cleaned_sales_data.csv and produces:
  - sales_report.xlsx   (multi-sheet Excel workbook with summary tables)
  - charts/             (PNG charts saved separately)

Run AFTER clean_data.py.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import os
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Helpers ───────────────────────────────────────────────────────────────────
BLUE_DARK  = "1F3864"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GREY_LIGHT = "F2F2F2"
GREEN      = "375623"
WHITE      = "FFFFFF"

def hdr_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def hdr_font(hex_color=WHITE, bold=True, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Arial")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def autofit(ws, extra=4):
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + extra

def write_table(ws, df, start_row=1, start_col=1,
                hdr_bg=BLUE_MID, hdr_fg=WHITE, number_cols=None):
    number_cols = number_cols or []
    # Header
    for ci, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font      = hdr_font(hdr_fg)
        cell.fill      = hdr_fill(hdr_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for ci, val in enumerate(row, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = hdr_fill(GREY_LIGHT if ri % 2 == 0 else WHITE)
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="right" if ci - start_col + 1 in number_cols
                                       else "left", vertical="center")
            if ci - start_col + 1 in number_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
    return ri  # last row written

# ── Load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv("cleaned_sales_data.csv", parse_dates=["order_date"])

os.makedirs("charts", exist_ok=True)
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "Executive Summary"
ws_sum.sheet_view.showGridLines = False

# Title banner
ws_sum.merge_cells("A1:F1")
title_cell = ws_sum["A1"]
title_cell.value = "Sales Performance Report – Automated Summary"
title_cell.font  = Font(name="Arial", size=16, bold=True, color=WHITE)
title_cell.fill  = hdr_fill(BLUE_DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 36

# KPI cards (row 3)
total_revenue  = df["revenue"].sum()
total_orders   = len(df)
avg_order_val  = df["revenue"].mean()
completed_pct  = (df["status"] == "Completed").mean() * 100

kpis = [
    ("Total Revenue",    f"₹{total_revenue:,.0f}"),
    ("Total Orders",     f"{total_orders:,}"),
    ("Avg Order Value",  f"₹{avg_order_val:,.2f}"),
    ("Completion Rate",  f"{completed_pct:.1f}%"),
]

ws_sum.row_dimensions[3].height = 14
ws_sum.row_dimensions[4].height = 30
ws_sum.row_dimensions[5].height = 22

for ci, (label, value) in enumerate(kpis, start=1):
    lc = ws_sum.cell(row=3, column=ci, value=label)
    lc.font = Font(name="Arial", size=9, color="595959")
    lc.alignment = Alignment(horizontal="center")

    vc = ws_sum.cell(row=4, column=ci, value=value)
    vc.font  = Font(name="Arial", size=14, bold=True, color=BLUE_DARK)
    vc.fill  = hdr_fill(BLUE_LIGHT)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = thin_border()

    ws_sum.column_dimensions[get_column_letter(ci)].width = 22


# ── Revenue by Region table ───────────────────────────────────────────────────
region_tbl = (df.groupby("region")["revenue"]
                .agg(Orders="count", Total_Revenue="sum", Avg_Revenue="mean")
                .reset_index()
                .sort_values("Total_Revenue", ascending=False))
region_tbl.columns = ["Region", "Orders", "Total Revenue (₹)", "Avg Revenue (₹)"]

ws_sum.cell(row=7, column=1, value="Revenue by Region").font = Font(
    name="Arial", size=11, bold=True, color=BLUE_DARK)
write_table(ws_sum, region_tbl, start_row=8, number_cols=[2, 3, 4])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – MONTHLY TRENDS
# ═══════════════════════════════════════════════════════════════════════════════
ws_mo = wb.create_sheet("Monthly Trends")
ws_mo.sheet_view.showGridLines = False

monthly = (df.groupby("order_month")["revenue"]
             .agg(Orders="count", Revenue="sum")
             .reset_index())
monthly.columns = ["Month", "Orders", "Revenue (₹)"]

ws_mo.cell(row=1, column=1, value="Monthly Sales Trends").font = Font(
    name="Arial", size=13, bold=True, color=BLUE_DARK)
write_table(ws_mo, monthly, start_row=2, number_cols=[2, 3])
autofit(ws_mo)

# Chart – monthly revenue
fig, ax = plt.subplots(figsize=(10, 4))
ax.bar(monthly["Month"], monthly["Revenue (₹)"] / 1000,
       color="#2E75B6", edgecolor="white", linewidth=0.5)
ax.set_title("Monthly Revenue (₹ Thousands)", fontsize=13, fontweight="bold", pad=12)
ax.set_xlabel("Month")
ax.set_ylabel("Revenue (₹K)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"₹{x:,.0f}K"))
plt.xticks(rotation=45, ha="right", fontsize=8)
plt.tight_layout()
chart_path = "charts/monthly_revenue.png"
plt.savefig(chart_path, dpi=150)
plt.close()

img = XLImage(chart_path)
img.anchor = f"E2"
ws_mo.add_image(img)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – CATEGORY ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws_cat = wb.create_sheet("Category Analysis")
ws_cat.sheet_view.showGridLines = False

cat_tbl = (df.groupby("category")
             .agg(Orders=("order_id", "count"),
                  Revenue=("revenue", "sum"),
                  Avg_Price=("unit_price", "mean"),
                  Avg_Qty=("quantity", "mean"))
             .reset_index()
             .sort_values("Revenue", ascending=False))
cat_tbl.columns = ["Category", "Orders", "Revenue (₹)", "Avg Unit Price (₹)", "Avg Qty"]
cat_tbl["Revenue (₹)"] = cat_tbl["Revenue (₹)"].round(2)
cat_tbl["Avg Unit Price (₹)"] = cat_tbl["Avg Unit Price (₹)"].round(2)
cat_tbl["Avg Qty"] = cat_tbl["Avg Qty"].round(1)

ws_cat.cell(row=1, column=1, value="Sales by Category").font = Font(
    name="Arial", size=13, bold=True, color=BLUE_DARK)
write_table(ws_cat, cat_tbl, start_row=2, number_cols=[2, 3, 4, 5])
autofit(ws_cat)

# Pie chart
fig, ax = plt.subplots(figsize=(7, 5))
colors = ["#2E75B6", "#70AD47", "#ED7D31", "#FFC000", "#A5A5A5", "#5A5A5A"]
wedges, texts, autotexts = ax.pie(
    cat_tbl["Revenue (₹)"], labels=cat_tbl["Category"],
    autopct="%1.1f%%", colors=colors[:len(cat_tbl)],
    startangle=140, pctdistance=0.82)
for t in autotexts:
    t.set_fontsize(9)
ax.set_title("Revenue Share by Category", fontsize=13, fontweight="bold")
plt.tight_layout()
pie_path = "charts/category_pie.png"
plt.savefig(pie_path, dpi=150)
plt.close()

img2 = XLImage(pie_path)
img2.anchor = "G2"
ws_cat.add_image(img2)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – STATUS BREAKDOWN
# ═══════════════════════════════════════════════════════════════════════════════
ws_st = wb.create_sheet("Status Breakdown")
ws_st.sheet_view.showGridLines = False

status_tbl = (df.groupby("status")
                .agg(Orders=("order_id", "count"),
                     Revenue=("revenue", "sum"))
                .reset_index()
                .sort_values("Orders", ascending=False))
status_tbl["Revenue (₹)"] = status_tbl["Revenue"].round(2)
status_tbl["Share (%)"]   = (status_tbl["Orders"] / status_tbl["Orders"].sum() * 100).round(1)
status_tbl = status_tbl[["status", "Orders", "Revenue (₹)", "Share (%)"]]
status_tbl.columns = ["Status", "Orders", "Revenue (₹)", "Share (%)"]

ws_st.cell(row=1, column=1, value="Order Status Breakdown").font = Font(
    name="Arial", size=13, bold=True, color=BLUE_DARK)
write_table(ws_st, status_tbl, start_row=2, number_cols=[2, 3, 4])
autofit(ws_st)

# Horizontal bar chart
fig, ax = plt.subplots(figsize=(7, 3.5))
colors_st = {"Completed": "#70AD47", "Pending": "#FFC000", "Cancelled": "#FF0000",
             "Unknown": "#A5A5A5"}
bar_colors = [colors_st.get(s, "#2E75B6") for s in status_tbl["Status"]]
ax.barh(status_tbl["Status"], status_tbl["Orders"], color=bar_colors, edgecolor="white")
ax.set_title("Orders by Status", fontsize=12, fontweight="bold")
ax.set_xlabel("Number of Orders")
ax.invert_yaxis()
plt.tight_layout()
status_path = "charts/status_breakdown.png"
plt.savefig(status_path, dpi=150)
plt.close()

img3 = XLImage(status_path)
img3.anchor = "F2"
ws_st.add_image(img3)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – RAW CLEANED DATA (first 200 rows)
# ═══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.create_sheet("Cleaned Data (Sample)")
ws_raw.sheet_view.showGridLines = False

sample = df.head(200).copy()
sample["order_date"] = sample["order_date"].astype(str)

ws_raw.cell(row=1, column=1,
            value=f"Cleaned Dataset Sample (first 200 of {len(df)} rows)").font = Font(
    name="Arial", size=11, bold=True, color=BLUE_DARK)
write_table(ws_raw, sample, start_row=2)
autofit(ws_raw)


# ── Save workbook ─────────────────────────────────────────────────────────────
wb.save("sales_report.xlsx")
print("✅  sales_report.xlsx created (5 sheets)")
print("✅  Charts saved in charts/")
print(f"\n📊 Quick stats:")
print(f"   Total Revenue : ₹{total_revenue:,.2f}")
print(f"   Total Orders  : {total_orders}")
print(f"   Avg Order Val : ₹{avg_order_val:,.2f}")
print(f"   Completion %  : {completed_pct:.1f}%")
